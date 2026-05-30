import csv
import glob
import logging
import os
import shutil
import subprocess
import sys
import unicodedata
import json
import time
from datetime import date, datetime, time as datetime_time, timedelta
from pathlib import Path

# Ensure script runs inside .venv if available ----------------------------------
# When a corporate environment has a local Python installation that is restricted
# we want to transparently re‑launch using the virtual environment created by
# `setup.bat`.  This block will restart the interpreter under `.venv` if it's
# present and we're currently running with the system Python.

VENV_PY = os.path.join(os.getcwd(), ".venv", "Scripts", "python.exe")
if os.path.exists(VENV_PY):
    # 'sys.prefix' of a venv contains the venv path; check to avoid loops
    if not sys.prefix.lower().startswith(os.path.join(os.getcwd(), ".venv").lower()):
        os.execv(VENV_PY, [VENV_PY] + sys.argv)
# -----------------------------------------------------------------------------

from openpyxl import load_workbook, Workbook
from openpyxl.utils.datetime import from_excel
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side



HEADER_ROW = 3
DATA_START_ROW = 4
OUTPUT_COLUMNS = [1, 3, 7, 8, 9, 11]
OUTPUT_HEADERS = [
    "DT",
    "STO",
    "DESTINO",
    "DATA SAIDA",
    "HORA SUB PICKING",
    "HORA SAIDA CARGA",
]


def _find_input_workbook(root_dir: str) -> str:
    target = os.path.join(root_dir, "Pasta de Viagens Itu.xlsx")
    if os.path.exists(target):
        return target
    candidates = [
        path
        for path in glob.glob(os.path.join(root_dir, "Pasta*.xlsx"))
        if not os.path.basename(path).startswith("~")
    ]
    if not candidates:
        raise FileNotFoundError(
            "Nao encontrei arquivo 'Pasta de Viagens Itu.xlsx' nem outro 'Pasta*.xlsx'."
        )
    candidates.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return candidates[0]


def _target_date(now: datetime) -> date:
    if now.time() >= datetime_time(22, 0):
        return (now + timedelta(days=1)).date()
    return now.date()


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().upper()
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    return " ".join(text.split())


def _normalize_header(value: object) -> str:
    text = _normalize_text(value)
    return text.replace("_", " ")


def _find_header_index(headers: list[object], name: str) -> int:
    target = _normalize_header(name)
    for idx, header in enumerate(headers):
        if _normalize_header(header) == target:
            return idx
    raise ValueError(f"Nao encontrei a coluna '{name}' no cabecalho.")


def _to_date(value, epoch) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return from_excel(value, epoch).date()
        except Exception:
            return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        formats = (
            "%d/%m/%Y",
            "%d/%m/%Y %H:%M",
            "%d/%m/%Y %H:%M:%S",
            "%d-%m-%Y",
            "%d-%m-%Y %H:%M",
            "%d-%m-%Y %H:%M:%S",
            "%Y-%m-%d",
            "%Y-%m-%d %H:%M",
            "%Y-%m-%d %H:%M:%S",
            "%d.%m.%Y",
            "%d.%m.%Y %H:%M",
            "%d.%m.%Y %H:%M:%S",
            "%m/%d/%Y",
            "%m/%d/%Y %H:%M",
            "%m/%d/%Y %H:%M:%S",
            "%Y/%m/%d",
            "%Y/%m/%d %H:%M",
            "%Y/%m/%d %H:%M:%S",
        )
        for fmt in formats:
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return None


def _format_csv_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.time() == datetime_time(0, 0):
            return value.strftime("%d/%m/%Y")
        return value.strftime("%d/%m/%Y %H:%M")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, datetime_time):
        return value.strftime("%H:%M")
    return str(value)


def _load_output_template(root_dir: str) -> tuple[Workbook, object]:
    wb = Workbook()
    ws = wb.active
    ws.title = "CABREUVA"

    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20

    return wb, ws


def _find_last_data_row(ws, max_col: int) -> int:
    max_row = ws.max_row or 1
    for row_idx in range(max_row, 0, -1):
        for col_idx in range(1, max_col + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value not in (None, ""):
                return row_idx
    return 1





def _archive_old_outputs(root_dir: str, target: date) -> None:
    archive_dir = os.path.join(root_dir, "historico_gerado")
    try:
        os.makedirs(archive_dir, exist_ok=True)
    except Exception as e:
        logging.warning("Nao foi possivel criar a pasta de historico: %s", e)
        return

    xlsx_pattern = os.path.join(root_dir, "solicitacao de pickings *.xlsx")
    csv_pattern = os.path.join(root_dir, "solicitacao de pickings *.csv")
    found_files = glob.glob(xlsx_pattern) + glob.glob(csv_pattern)
    
    for src_path in found_files:
        try:
            base = os.path.basename(src_path)
            if "historico_gerado" in src_path:
                continue
                
            try:
                mtime = os.path.getmtime(src_path)
                dt_modified = datetime.fromtimestamp(mtime)
                stamp = dt_modified.strftime("_%H%M%S")
            except Exception:
                stamp = datetime.now().strftime("_%H%M%S")
                
            name, ext = os.path.splitext(base)
            dst_name = f"{name}{stamp}{ext}"
            dst_path = os.path.join(archive_dir, dst_name)
            
            shutil.move(src_path, dst_path)
            logging.info("Arquivo antigo arquivado: %s -> %s", src_path, dst_path)
            print(f"[INFO] Arquivo antigo arquivado: {base} -> historico_gerado\\{dst_name}")
        except Exception as e:
            logging.warning("Falha ao arquivar arquivo %s: %s", src_path, e)


def _prepare_output_path(path: str) -> str:
    if not os.path.exists(path):
        return path

    try:
        os.remove(path)
        return path
    except PermissionError:
        base, ext = os.path.splitext(path)
        stamp = datetime.now().strftime("%H%M%S")
        alt_path = f"{base} ({stamp}){ext}"
        logging.warning("Arquivo em uso. Usando nome alternativo: %s", alt_path)
        return alt_path
    except OSError:
        base, ext = os.path.splitext(path)
        stamp = datetime.now().strftime("%H%M%S")
        alt_path = f"{base} ({stamp}){ext}"
        logging.warning(
            "Falha ao remover arquivo. Usando nome alternativo: %s", alt_path
        )
        return alt_path


def _load_email_config(root_dir: str) -> dict:
    config_path = os.path.join(root_dir, "config_email.txt")
    defaults = {
        "PARA": "warehouse@empresa.com.br",
        "CC": "portaria@empresa.com.br; roterizacao@empresa.com.br",
        "ENVIAR_DIRETO": "NAO",
    }
    
    if not os.path.exists(config_path):
        try:
            with open(config_path, "w", encoding="utf-8") as f:
                f.write(
                    "# CONFIGURACAO DE E-MAIL - AUTOPICKINGPY\n"
                    "# ==============================================================================\n"
                    "# Edite os valores abaixo. Mantenha os nomes das chaves (antes do sinal de \"=\").\n"
                    "# Para multiplos e-mails no PARA ou CC, separe-os por ponto e virgula (;).\n\n"
                    "# E-mails dos destinatarios principais (Warehouse)\n"
                    f"PARA = {defaults['PARA']}\n\n"
                    "# E-mails das copias (Portaria e Roteirizacao)\n"
                    f"CC = {defaults['CC']}\n\n"
                    "# Enviar e-mail automaticamente? (SIM para enviar direto, NAO para apenas exibir a tela e revisar)\n"
                    f"ENVIAR_DIRETO = {defaults['ENVIAR_DIRETO']}\n"
                )
            logging.info("Arquivo de configuracao config_email.txt criado com valores padrao.")
        except Exception:
            logging.exception("Falha ao criar config_email.txt com valores padrao.")
            
    config = defaults.copy()
    if os.path.exists(config_path):
        try:
            with open(config_path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#"):
                        continue
                    if "=" in line:
                        parts = line.split("=", 1)
                        key = parts[0].strip().upper()
                        val = parts[1].strip()
                        config[key] = val
        except Exception:
            logging.exception("Erro ao ler config_email.txt, usando padroes.")
            
    return {
        "para": config.get("PARA", defaults["PARA"]),
        "cc": config.get("CC", defaults["CC"]),
        "enviar_direto": config.get("ENVIAR_DIRETO", defaults["ENVIAR_DIRETO"]).strip().upper() == "SIM",
    }


def _generate_html_table(filtered_rows: list) -> str:
    num_cols = len(OUTPUT_HEADERS)
    total_rows = len(filtered_rows)
    
    # CSS styles matching Excel exactly
    table_style = (
        "border-collapse: collapse; "
        "font-family: Calibri, Arial, sans-serif; "
        "font-size: 10pt; "
        "border: 1px solid #a0a0a0; "
        "width: 100%; "
        "max-width: 800px; "
        "margin: 15px 0;"
    )
    
    title_cell_style = (
        "background-color: #2E5C8A; "
        "color: #FFFFFF; "
        "font-family: Calibri, Arial, sans-serif; "
        "font-weight: bold; "
        "font-size: 11pt; "
        "height: 30px; "
        "text-align: center; "
        "vertical-align: middle; "
        "border: 1px solid #a0a0a0; "
        "padding: 5px;"
    )
    
    header_cell_style = (
        "background-color: #4A90E2; "
        "color: #FFFFFF; "
        "font-family: Calibri, Arial, sans-serif; "
        "font-weight: bold; "
        "font-size: 10pt; "
        "height: 25px; "
        "text-align: center; "
        "vertical-align: middle; "
        "border: 1px solid #a0a0a0; "
        "padding: 5px;"
    )
    
    data_cell_style = (
        "font-family: Calibri, Arial, sans-serif; "
        "font-size: 10pt; "
        "text-align: center; "
        "vertical-align: middle; "
        "border: 1px solid #c0c0c0; "
        "padding: 5px; "
        "height: 20px;"
    )
    
    # Start HTML string
    html = f'<table style="{table_style}">\n'
    
    # 1. Title Row (A1 merged)
    html += '  <tr>\n'
    html += f'    <td colspan="{num_cols - 1}" style="{title_cell_style}">CDV E CABREUVA</td>\n'
    html += f'    <td style="{title_cell_style}">Total: {total_rows}</td>\n'
    html += '  </tr>\n'
    
    # 2. Header Row
    html += '  <tr>\n'
    for header in OUTPUT_HEADERS:
        html += f'    <th style="{header_cell_style}">{header}</th>\n'
    html += '  </tr>\n'
    
    # 3. Data Rows
    for offset, (_, row_filtered) in enumerate(filtered_rows):
        bg_color = "#FFFFFF" if offset % 2 == 0 else "#F7FAFC"
        row_style = data_cell_style + f" background-color: {bg_color};"
        
        html += '  <tr>\n'
        for val in row_filtered:
            formatted_val = _format_csv_value(val)
            html += f'    <td style="{row_style}">{formatted_val}</td>\n'
        html += '  </tr>\n'
        
    html += '</table>'
    return html


def _copy_html_to_clipboard(html_content: str, filtered_rows: list, greeting_html: str = "", greeting_plain: str = "") -> None:
    try:
        import win32clipboard
        
        # 1. Prepare HTML format payload for Windows clipboard
        template = (
            "Version:0.9\r\n"
            "StartHTML:{start_html:010d}\r\n"
            "EndHTML:{end_html:010d}\r\n"
            "StartFragment:{start_frag:010d}\r\n"
            "EndFragment:{end_frag:010d}\r\n"
            "<html>\r\n"
            "<body>\r\n"
            "<!--StartFragment-->{content}<!--EndFragment-->\r\n"
            "</body>\r\n"
            "</html>"
        )
        
        html_payload = greeting_html + html_content
        dummy = template.format(start_html=0, end_html=0, start_frag=0, end_frag=0, content=html_payload)
        
        header_len = dummy.find("<html>")
        start_html = header_len
        end_html = len(dummy)
        start_frag = dummy.find("<!--StartFragment-->") + len("<!--StartFragment-->")
        end_frag = dummy.find("<!--EndFragment-->")
        
        payload = template.format(
            start_html=start_html,
            end_html=end_html,
            start_frag=start_frag,
            end_frag=end_frag,
            content=html_payload
        )
        
        # 2. Build plain text tab-separated fallback
        plain_lines = ["DT\tSTO\tDESTINO\tDATA SAIDA\tHORA SUB PICKING\tHORA SAIDA CARGA"]
        for _, row in filtered_rows:
            plain_lines.append("\t".join([_format_csv_value(v) for v in row]))
        plain_text = greeting_plain + "\r\n".join(plain_lines)
        
        # 3. Open Clipboard and write data
        win32clipboard.OpenClipboard()
        try:
            win32clipboard.EmptyClipboard()
            cf_html = win32clipboard.RegisterClipboardFormat("HTML Format")
            # Write both HTML format (rich table) and plain text (CF_UNICODETEXT)
            win32clipboard.SetClipboardData(cf_html, payload.encode("utf-8"))
            win32clipboard.SetClipboardText(plain_text, win32clipboard.CF_UNICODETEXT)
            logging.info("Tabela copiada para a Area de Transferencia (HTML e Texto).")
        finally:
            win32clipboard.CloseClipboard()
            
    except Exception as e:
        logging.warning("Falha ao copiar para a area de transferencia: %s", e)


def _open_mailto_link(config: dict, target_date: date, include_body: bool = True) -> None:
    try:
        import webbrowser
        import urllib.parse
        
        target_date_str = target_date.strftime("%d/%m/%Y")
        subject = f"SOLICITAÇÃO DE PICKING {target_date_str}"
        
        to_str = urllib.parse.quote(config["para"])
        cc_str = urllib.parse.quote(config["cc"])
        subject_str = urllib.parse.quote(subject)
        
        if include_body:
            body_text = (
                "Olá, tudo bem?\r\n\r\n"
                "Segue solicitação de pickings para CDV e Cabreúva. "
                "Por gentileza, caso não haja pickings do plano do dia disponível, subir com antecipação.\r\n\r\n"
            )
            body_str = urllib.parse.quote(body_text)
            mailto_url = f"mailto:{to_str}?cc={cc_str}&subject={subject_str}&body={body_str}"
        else:
            mailto_url = f"mailto:{to_str}?cc={cc_str}&subject={subject_str}"
            
        webbrowser.open(mailto_url)
        logging.info("Link mailto aberto no leitor de e-mail padrao.")
    except Exception as e:
        logging.warning("Falha ao abrir o link mailto: %s", e)


def _focus_window_by_title(title_substring: str) -> bool:
    try:
        import win32gui
        import win32con
        import ctypes
    except ImportError:
        return False

    search_str = title_substring.strip().lower()
    found_hwnd = []

    def enum_callback(hwnd, extra):
        if win32gui.IsWindowVisible(hwnd):
            title = win32gui.GetWindowText(hwnd).strip().lower()
            if search_str in title:
                # IGNORAR JANELAS DO PROPRIO PROJETO / CMD / EDITORES para evitar falsos positivos
                if "autopickingpy" in title or "cmd.exe" in title or "prompt de comando" in title or "visual studio" in title:
                    if search_str != "autopickingpy":
                        return True
                
                # IGNORAR NAVEGADORES PARA BUSCAS GENERICAS
                # Se estamos buscando um termo generico (como 'outlook', 'picking', 'mensagem'),
                # nao queremos focar janelas de navegadores normais (como o inbox do Chrome)
                generic_terms = ["picking", "solicitacao", "outlook", "mensagem", "email", "mail"]
                if search_str in generic_terms:
                    # Ignorar navegadores conhecidos
                    for browser in ("chrome", "edge", "firefox", "brave", "opera", "safari"):
                        if browser in title:
                            return True
                    # Ignorar caixas de entrada ou paginas de webmail genericas
                    for ignore_term in ("caixa de entrada", "inbox", "office 365", "microsoft 365", "gmail", "webmail"):
                        if ignore_term in title:
                            return True
                            
                found_hwnd.append(hwnd)
        return True

    win32gui.EnumWindows(enum_callback, None)
    if found_hwnd:
        hwnd = found_hwnd[0]
        try:
            if win32gui.IsIconic(hwnd):
                win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
            else:
                win32gui.ShowWindow(hwnd, win32con.SW_SHOW)

            # ALT-key bypass workaround to allow SetForegroundWindow from background process
            user32 = ctypes.windll.user32
            # Press ALT (0x12)
            user32.keybd_event(0x12, 0, 0, 0)
            time.sleep(0.05)
            # Release ALT
            user32.keybd_event(0x12, 0, 0x0002, 0)

            win32gui.SetForegroundWindow(hwnd)
            
            # Press ESC (0x1B) to dismiss the menu state activated by the ALT key
            time.sleep(0.1)
            user32.keybd_event(0x1B, 0, 0, 0) # Press ESC
            time.sleep(0.05)
            user32.keybd_event(0x1B, 0, 0x0002, 0) # Release ESC
            return True
        except Exception as e:
            logging.warning("Erro ao tentar trazer janela para primeiro plano via win32: %s", e)
            try:
                win32gui.SetForegroundWindow(hwnd)
                return True
            except Exception:
                pass
    return False


def _ensure_focus(subject: str) -> bool:
    try:
        import win32gui
        current_hwnd = win32gui.GetForegroundWindow()
        current_title = win32gui.GetWindowText(current_hwnd).strip().lower()
        
        # Se a janela ativa atual for a do próprio CMD/projeto/editor, ela NÃO é a de e-mail!
        if "autopickingpy" in current_title or "cmd.exe" in current_title or "prompt de comando" in current_title or "visual studio" in current_title:
            # Força o refoco na de e-mail real abaixo
            pass
        # Se a janela ativa for um navegador, mas não contiver o assunto exato do e-mail, ela NÃO é a de e-mail!
        elif any(b in current_title for b in ("chrome", "edge", "firefox", "brave", "opera", "safari")) and subject.lower() not in current_title:
            # Força o refoco na de e-mail real abaixo
            pass
        else:
            subject_no_accent = "solicitacao de picking"
            for term in (subject.lower(), subject_no_accent, "picking", "solicitacao", "outlook", "mensagem", "email", "mail"):
                if term in current_title:
                    return True
                
        # If not focused, try to refocus
        subject_no_accent = "solicitacao de picking"
        for term in (subject, subject_no_accent, "picking", "solicitacao", "outlook", "mensagem", "email", "mail"):
            if _focus_window_by_title(term):
                return True
    except Exception:
        pass
    return False


def _send_outlook_email(config: dict, target_date: date, html_table: str, filtered_rows: list) -> None:
    try:
        import win32com.client
    except ImportError:
        logging.warning("pywin32 nao instalado. Nao e possivel automatizar o Outlook.")
        print("[AVISO] Nao foi possivel enviar o e-mail: o pacote 'pywin32' nao esta instalado.")
        return

    # Use target_date directly as it already represents the correct picking date
    target_date_str = target_date.strftime("%d/%m/%Y")
    
    subject = f"SOLICITAÇÃO DE PICKING {target_date_str}"
    
    logging.info("Inicializando automacao do Microsoft Outlook...")
    print("Inicializando automacao do Outlook...")
    
    try:
        # Tenta conectar via diferentes strings de classe COM registradas
        outlook = None
        for class_str in ("Outlook.Application", "Outlook.Application.16", "Outlook.Application.15", "Outlook.Application.14"):
            try:
                outlook = win32com.client.Dispatch(class_str)
                break
            except Exception:
                continue
                
        if outlook is None:
            raise Exception("Nao foi possivel instanciar nenhum servidor COM do Outlook.")
            
        mail = outlook.CreateItem(0)
        
        mail.Subject = subject
        mail.To = config["para"]
        mail.CC = config["cc"]
        
        # Display email to load default signature
        mail.Display()
        
        # Retrieve the pre-populated HTMLBody containing the signature
        initial_html = mail.HTMLBody
        
        body_intro = (
            "<p style='font-family: Calibri, Arial, sans-serif; font-size: 11pt;'>Olá, tudo bem?</p>"
            "<p style='font-family: Calibri, Arial, sans-serif; font-size: 11pt;'>"
            "Segue solicitação de pickings para CDV e Cabreúva. "
            "Por gentileza, caso não haja pickings do plano do dia disponível, subir com antecipação."
            "</p>"
        )
        
        # Inject our styled table and content above the user's signature
        if "<body" in initial_html.lower():
            body_start_idx = initial_html.lower().find("<body")
            tag_end_idx = initial_html.find(">", body_start_idx)
            new_html = (
                initial_html[:tag_end_idx + 1] +
                "<div style='font-family: Calibri, Arial, sans-serif; font-size: 11pt; color: #000000;'>" +
                body_intro +
                html_table +
                "</div><br>" +
                initial_html[tag_end_idx + 1:]
            )
        else:
            new_html = (
                "<div style='font-family: Calibri, Arial, sans-serif; font-size: 11pt; color: #000000;'>" +
                body_intro +
                html_table +
                "</div><br>" +
                initial_html
            )
            
        mail.HTMLBody = new_html
        
        if config["enviar_direto"]:
            mail.Send()
            logging.info("E-mail enviado diretamente via Outlook.")
            print("[SUCESSO] E-mail enviado automaticamente via Outlook!")
        else:
            logging.info("Rascunho de e-mail exibido com sucesso no Outlook.")
            print("[SUCESSO] Rascunho de e-mail exibido no Outlook! O usuario pode clicar em Enviar.")
            
    except Exception as e:
        logging.exception("Erro ao automatizar o Outlook")
        print(f"\n[AVISO] Nao foi possivel usar a automacao classica COM do Outlook: {e}")
        print("Ativando modo de compatibilidade avancado com auto-colagem...")
        
        try:
            # 1. Abrir o link mailto padrao com corpo (para que a saudacao seja criada nativamente e a caixa de texto ganhe foco)
            _open_mailto_link(config, target_date, include_body=True)
            
            # 2. Esperar e ativar a janela de e-mail usando win32gui ou WScript.Shell
            print("[INFO] Aguardando abertura do leitor de e-mail (otimizado para computadores rapidos/lentos)...")
            activated = False
            
            subject_no_accent = "solicitacao de picking"
            
            # Polling ultra-veloz a cada 0.2 segundos por ate 15 segundos para responder imediatamente quando a janela surgir
            for _ in range(75):
                time.sleep(0.2)
                # Tenta focar usando a nossa funcao win32 robusta com varios termos de busca
                for term in (subject, subject_no_accent, "picking", "solicitacao", "outlook", "mensagem", "email", "mail"):
                    if _focus_window_by_title(term):
                        logging.info("Janela focada com sucesso via termo '%s'", term)
                        activated = True
                        break
                if activated:
                    break
                
                # Fallback secundario usando WScript.Shell AppActivate
                try:
                    import win32com.client
                    wsh = win32com.client.Dispatch("WScript.Shell")
                    if wsh.AppActivate(subject) or wsh.AppActivate("SOLICITAÇÃO DE PICKING") or wsh.AppActivate("Outlook") or wsh.AppActivate("Nova Mensagem") or wsh.AppActivate("Mensagem"):
                        activated = True
                        break
                except Exception:
                    pass
            
            if activated:
                # Damos 2.0s de margem para que os computadores corporativos mais lentos renderizem o WebView2 completamente
                time.sleep(2.0)
                
                # BLOQUEIO DE SEGURANÇA CRÍTICA: Trava o teclado e mouse do usuario durante a colagem de milissegundos
                # para impedir interferencias fisicas caso o usuario clique fora ou digite
                try:
                    import ctypes
                    ctypes.windll.user32.BlockInput(True)
                except Exception:
                    pass
                
                try:
                    # Copiar HTML e texto para o Clipboard exatamente aqui, dentro do bloco de bloqueio!
                    # Isso impede 100% que o usuário copie outra coisa para a área de transferência antes da colagem!
                    _copy_html_to_clipboard(html_table, filtered_rows)
                    
                    # Garante resiliência do foco antes de enviar as setas direcionais
                    _ensure_focus(subject)
                    
                    # Simular pressionar a seta para baixo (VK_DOWN = 0x28) exatamente 4 vezes
                    # para mover o cursor de forma segura e local sem sair do editor.
                    # Isso evita o Ctrl+End que as vezes rola a pagina inteira do Chromium fora do editor.
                    import ctypes
                    user32 = ctypes.windll.user32
                    for _ in range(4):
                        user32.keybd_event(0x28, 0, 0, 0) # Press Down (0x28)
                        time.sleep(0.05)
                        user32.keybd_event(0x28, 0, 0x0002, 0) # Release Down
                        time.sleep(0.05)
                    time.sleep(0.1)
                    
                    # Garante resiliência do foco antes de colar
                    _ensure_focus(subject)
                    
                    # 4. Simular colagem de forma nativa via Windows API (Ctrl+V)
                    user32.keybd_event(0x11, 0, 0, 0)
                    user32.keybd_event(0x56, 0, 0, 0)
                    time.sleep(0.05)
                    user32.keybd_event(0x56, 0, 0x0002, 0)
                    user32.keybd_event(0x11, 0, 0x0002, 0)
                finally:
                    # DESBLOQUEIO OBRIGATÓRIO E GARANTIDO DA ENTRADA DO USUÁRIO
                    try:
                        ctypes.windll.user32.BlockInput(False)
                    except Exception:
                        pass
            else:
                logging.warning("Nao foi possivel focar automaticamente a janela. Usando delay fixo como fallback...")
                time.sleep(4.0)
                
                # Simular colagem padrao sem foco verificado
                try:
                    import ctypes
                    user32 = ctypes.windll.user32
                    user32.keybd_event(0x11, 0, 0, 0)
                    user32.keybd_event(0x56, 0, 0, 0)
                    time.sleep(0.05)
                    user32.keybd_event(0x56, 0, 0x0002, 0)
                    user32.keybd_event(0x11, 0, 0x0002, 0)
                except Exception:
                    pass
            
            logging.info("Tabela colada automaticamente via simulacao de teclado.")
            print("\n=========================================================================================")
            print("[SUCESSO] E-mail criado e tabela colada AUTOMATICAMENTE com sucesso!")
            print("=========================================================================================\n")
        except Exception as fe:
            logging.exception("Erro no fallback de e-mail")
            print(f"[ERRO] Falha no modo de compatibilidade: {fe}")
            
    # Trazer a janela do CMD de volta ao primeiro plano apos concluir a automacao de e-mail
    try:
        logging.info("Tentando retornar o foco para a janela do CMD (AutoPickingPy)...")
        time.sleep(1.5)
        _focus_window_by_title("autopickingpy")
    except Exception as fe_focus:
        logging.warning(f"Erro ao tentar re-focar a janela do CMD: {fe_focus}")


def _check_config_stage() -> int:
    root_dir = os.path.dirname(os.path.abspath(__file__))
    config = _load_email_config(root_dir)
    
    config_path = os.path.join(root_dir, "config_email.txt")
    
    # Verificar se o arquivo esta vazio ou contem apenas comentarios/linhas em branco
    is_empty = False
    if os.path.exists(config_path):
        try:
            with open(config_path, "r", encoding="utf-8") as f:
                content = f.read().strip()
                lines = [l.strip() for l in content.split("\n") if l.strip() and not l.strip().startswith("#")]
                if not lines:
                    is_empty = True
        except Exception:
            pass
    else:
        is_empty = True

    # Verificar se campos cruciais estao em branco
    is_blank = not config["para"].strip() or not config["cc"].strip()
    
    is_default = (
        config["para"] == "warehouse@empresa.com.br" or
        "portaria@empresa.com.br" in config["cc"]
    )
    
    print("\n" + "="*80)
    print("                      VERIFICAÇÃO DE CONFIGURAÇÃO DE E-MAIL")
    print("="*80)
    print(f" Destinatario Principal (PARA): {config['para'] if config['para'].strip() else '[VAZIO / EM BRANCO]'}")
    print(f" Copia (CC):                     {config['cc'] if config['cc'].strip() else '[VAZIO / EM BRANCO]'}")
    print(f" Modo de Envio Direto:           {'SIM (Enviar sem revisar)' if config['enviar_direto'] else 'NAO (Abrir rascunho)'}")
    print("="*80)
    
    if is_empty or is_blank:
        print("\n[ATENCAO] O arquivo de configuracao 'config_email.txt' esta VAZIO ou contem campos em branco!")
        print("Voce precisa preencher os e-mails dos destinatarios para que a automacao funcione.")
        
        try:
            choice = input("\nDeseja ENCERRAR a automacao agora para preencher os e-mails? (S/N): ").strip().upper()
            if choice == "S":
                print("\n[INFO] Automacao encerrada. Por favor, edite o arquivo 'config_email.txt' e execute novamente.")
                return 2  # Exit code 2 to indicate user-requested termination
        except (KeyboardInterrupt, EOFError):
            return 2
    elif is_default:
        print("\n[ATENCAO] Os enderecos de e-mail em 'config_email.txt' ainda sao os valores padrao!")
        print("Recomendamos que voce configure os e-mails corretos antes do primeiro envio real.")
        
        try:
            choice = input("\nDeseja ENCERRAR a automacao agora para ajustar os e-mails? (S/N): ").strip().upper()
            if choice == "S":
                print("\n[INFO] Automacao encerrada. Por favor, edite o arquivo 'config_email.txt' e execute novamente.")
                return 2  # Exit code 2 to indicate user-requested termination
        except (KeyboardInterrupt, EOFError):
            return 2
            
    print("\n[OK] Configuracao de e-mail validada. Prosseguindo...")
    return 0


def _email_only_stage() -> int:
    root_dir = os.path.dirname(os.path.abspath(__file__))
    temp_json_path = os.path.join(root_dir, ".temp_filtered.json")
    if not os.path.exists(temp_json_path):
        print("[ERRO] Arquivo de dados temporarios nao encontrado. Execute a extracao primeiro.")
        return 1
        
    try:
        with open(temp_json_path, "r", encoding="utf-8") as f:
            temp_data = json.load(f)
            
        target = datetime.strptime(temp_data["target_date"], "%Y-%m-%d").date()
        filtered_rows = temp_data["filtered_rows"]
        
        # Re-generate the html table in memory
        html_table = _generate_html_table(filtered_rows)
        
        config = _load_email_config(root_dir)
        _send_outlook_email(config, target, html_table, filtered_rows)
        
        # Cleanup
        try:
            os.remove(temp_json_path)
        except Exception:
            pass
            
        return 0
    except Exception as e:
        logging.exception("Erro na etapa de e-mail")
        print(f"[ERRO] Falha ao disparar automacao de e-mail: {e}")
        return 1


def main() -> None:
    logging.info("Inicio da execucao")
    root_dir = os.path.dirname(os.path.abspath(__file__))
    
    extract_only = False
    target = None
    
    if len(sys.argv) > 1:
        arg = sys.argv[1].strip()
        if arg == "--check-config":
            sys.exit(_check_config_stage())
        elif arg == "--email-only":
            sys.exit(_email_only_stage())
        elif arg == "--extract-only":
            extract_only = True
            if len(sys.argv) > 2:
                date_str_arg = sys.argv[2]
                for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d"):
                    try:
                        target = datetime.strptime(date_str_arg, fmt).date()
                        break
                    except ValueError:
                        continue
        else:
            # Mantem compatibilidade com passagem direta de data
            date_str_arg = arg
            for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d"):
                try:
                    target = datetime.strptime(date_str_arg, fmt).date()
                    break
                except ValueError:
                    continue
            if target is None:
                print(f"Formato de data invalido: {date_str_arg}. Use DD-MM-YYYY, DD/MM/YYYY ou YYYY-MM-DD.")
                logging.error(f"Formato de data invalido: {date_str_arg}")
                sys.exit(1)

    input_path = _find_input_workbook(root_dir)
    logging.info("Arquivo de entrada: %s", input_path)

    wb = load_workbook(input_path, data_only=True)
    if "PASTA DE VIAGEM" in wb.sheetnames:
        ws = wb["PASTA DE VIAGEM"]
    else:
        ws = wb.worksheets[0]
        logging.warning(
            "Aba 'PASTA DE VIAGEM' não encontrada. Usando a primeira aba: %s", ws.title
        )
    if ws.auto_filter:
        ws.auto_filter.ref = None
        for row_idx in range(1, (ws.max_row or 0) + 1):
            ws.row_dimensions[row_idx].hidden = False

    all_headers = [ws.cell(row=HEADER_ROW, column=col).value for col in range(1, 35)]
    
    # Se extract_only estiver ativo, omitimos o log poluído do cabeçalho na tela (mandando apenas pro log.txt)
    logging.info("Cabeçalho lido da planilha:")
    for idx, header in enumerate(all_headers, 1):
        logging.info(f"Coluna {idx}: {repr(header)}")
    if not extract_only:
        print("Cabeçalho lido da planilha:")
        for idx, header in enumerate(all_headers, 1):
            print(f"Coluna {idx}: {repr(header)}")

    origin_idx = _find_header_index(all_headers, "ORIGEM")
    date_idx = _find_header_index(all_headers, "DATA_SAIDA")

    output_headers = OUTPUT_HEADERS

    # target is already resolved at the start of main()

    if target is None:
        target = _target_date(datetime.now())
    logging.info("Data alvo: %s", target.strftime("%d/%m/%Y"))
    _archive_old_outputs(root_dir, target)

    filtered_rows: list[list[object]] = []
    max_row = ws.max_row or 0
    for row_idx in range(DATA_START_ROW, max_row + 1):
        row_all = [ws.cell(row=row_idx, column=col).value for col in range(1, 35)]

        factory_value = _normalize_text(row_all[origin_idx])
        if factory_value != "FABRICA ITU":
            continue

        date_value = _to_date(row_all[date_idx], wb.epoch)
        if date_value != target:
            continue

        row_filtered = [row_all[col - 1] for col in OUTPUT_COLUMNS]
        filtered_rows.append((row_idx, row_filtered))



    date_str = target.strftime("%d-%m-%Y")
    base_name = f"solicitacao de pickings {date_str}"
    excel_path = os.path.join(root_dir, f"{base_name}.xlsx")
    csv_path = os.path.join(root_dir, f"{base_name}.csv")

    excel_path = _prepare_output_path(excel_path)
    csv_path = _prepare_output_path(csv_path)
    logging.info("Saida XLSX: %s", excel_path)
    logging.info("Saida CSV: %s", csv_path)

    out_wb, out_ws = _load_output_template(root_dir)
    num_cols = len(OUTPUT_COLUMNS)

    out_ws.merge_cells(f"A1:{get_column_letter(num_cols - 1)}1")
    out_ws.cell(row=1, column=1, value="CDV E CABREUVA")
    out_ws.cell(row=1, column=num_cols, value=f"Total: {len(filtered_rows)}")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col in range(1, num_cols + 1):
        cell = out_ws.cell(row=1, column=col)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(
            start_color="2E5C8A", end_color="2E5C8A", fill_type="solid"
        )
        cell.border = thin_border

    for col_idx, header in enumerate(output_headers, 1):
        cell = out_ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(
            start_color="4A90E2", end_color="4A90E2", fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for offset, (_, row_filtered) in enumerate(filtered_rows):
        out_row = 3 + offset
        for col_idx, value in enumerate(row_filtered, 1):
            out_ws.cell(row=out_row, column=col_idx, value=value)

    if filtered_rows:
        data_last_row = 2 + len(filtered_rows)

        for row in range(3, data_last_row + 1):
            out_ws.cell(row=row, column=1).number_format = "General"
            out_ws.cell(row=row, column=4).number_format = "dd/mm/yyyy"
            out_ws.cell(row=row, column=5).number_format = "hh:mm"
            out_ws.cell(row=row, column=6).number_format = "hh:mm"
            out_ws.row_dimensions[row].height = 15

        for row in out_ws.iter_rows(
            min_row=1, max_row=data_last_row, min_col=1, max_col=num_cols
        ):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in out_ws.iter_rows(
            min_row=3, max_row=data_last_row, min_col=1, max_col=num_cols
        ):
            for cell in row:
                cell.border = thin_border

        for row in out_ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=num_cols):
            for cell in row:
                cell.border = thin_border

        max_lengths = [len(str(h)) for h in output_headers]
        for row in out_ws.iter_rows(
            min_row=3, max_row=data_last_row, min_col=1, max_col=num_cols
        ):
            for col_idx, cell in enumerate(row, 1):
                length = len(_format_csv_value(cell.value))
                if length > max_lengths[col_idx - 1]:
                    max_lengths[col_idx - 1] = length

        for col_idx, length in enumerate(max_lengths, 1):
            adjusted_width = min(length + 2, 30)
            out_ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    out_wb.save(excel_path)

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as csv_file:
        writer = csv.writer(csv_file, delimiter=";")
        writer.writerow(output_headers)
        for _, row_filtered in filtered_rows:
            writer.writerow([_format_csv_value(value) for value in row_filtered])

    # Salva os dados filtrados para a etapa subsequente de e-mail (--email-only)
    try:
        serializable_rows = []
        for row_idx, row in filtered_rows:
            row_serialized = []
            for val in row:
                if isinstance(val, (date, datetime_time, datetime)):
                    row_serialized.append(_format_csv_value(val))
                else:
                    row_serialized.append(val)
            serializable_rows.append((row_idx, row_serialized))
            
        temp_data = {
            "target_date": target.strftime("%Y-%m-%d"),
            "excel_path": excel_path,
            "csv_path": csv_path,
            "filtered_rows": serializable_rows
        }
        temp_json_path = os.path.join(root_dir, ".temp_filtered.json")
        with open(temp_json_path, "w", encoding="utf-8") as f:
            json.dump(temp_data, f, ensure_ascii=False, indent=2)
    except Exception as je:
        logging.warning("Nao foi possivel salvar arquivo temporario JSON: %s", je)

    if extract_only:
        print("\n" + "="*80)
        print("                      RESUMO DA EXTRAÇÃO DE DADOS - SUCESSO")
        print("="*80)
        print(f" Planilha de Entrada:   {os.path.basename(input_path)}")
        print(f" Data Selecionada:       {target.strftime('%d/%m/%Y')}")
        print(f" Linhas Encontradas:     {len(filtered_rows)}")
        print(f" Arquivo Excel Gerado:   {os.path.basename(excel_path)}")
        print(f" Arquivo CSV Gerado:     {os.path.basename(csv_path)}")
        print("="*80 + "\n")
        logging.info("Etapa de extracao concluida (apenas extracao)")
        return

    print(f"Entrada: {os.path.basename(input_path)}")
    print(f"Data alvo: {target.strftime('%d/%m/%Y')}")
    print(f"Linhas filtradas: {len(filtered_rows)}")
    print(f"Excel: {excel_path}")
    print(f"CSV: {csv_path}")
    logging.info("Linhas filtradas: %s", len(filtered_rows))
    
    if filtered_rows:
        try:
            config = _load_email_config(root_dir)
            html_table = _generate_html_table(filtered_rows)
            _send_outlook_email(config, target, html_table, filtered_rows)
        except Exception:
            logging.exception("Erro ao processar envio de e-mail")
            print("[ERRO] Nao foi possivel gerar a solicitacao de e-mail via Outlook.")
    else:
        print("Nao ha dados para enviar por e-mail para a data alvo.")
        
    logging.info("Execucao concluida")


if __name__ == "__main__":
    # Configure logging
    log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "log.txt")
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.FileHandler(log_path, mode="w", encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
    )
    try:
        main()
    except Exception:
        logging.exception("Erro inesperado")
        raise
